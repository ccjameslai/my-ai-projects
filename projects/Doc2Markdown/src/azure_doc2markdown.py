import os
import re
import pandas as pd
from PyPDF2 import PdfReader, PdfWriter
from langchain_community.document_loaders import AzureAIDocumentIntelligenceLoader
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence.models import AnalyzeResult
from docx import Document
from PIL import Image
from azure.ai.documentintelligence.models import AnalyzeOutputOption, AnalyzeResult
import win32com.client
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import xlwings as xw
import time
from glob import glob
from pathlib import Path
from tqdm import tqdm
from dotenv import load_dotenv
import fitz 
import io

'''
Reference : https://learn.microsoft.com/en-us/python/api/overview/azure/ai-documentintelligence-readme?view=azure-python#using-prebuilt-models
'''

load_dotenv()

# 讀完.env 再印出
load_dotenv(override=True)

def Azure_Doc(file_path):
    
    # Linux
    endpoint = os.getenv("AZURE_AIDOC_ENDPOINT")
    key = os.getenv("AZURE_AIDOC_KEY")
    
    # Windows
    # endpoint = os.environ.get("AZURE_AIDOC_ENDPOINT")
    # key = os.environ.get("AZURE_AIDOC_KEY")
    
    loader = AzureAIDocumentIntelligenceLoader(
        api_endpoint=endpoint, api_key=key, file_path=file_path, api_model="prebuilt-layout", mode='markdown'
    )
    documents = loader.load()
    return documents

def parse_to_md(md_dir, file_path, image_paths):
    documents = Azure_Doc(file_path)
    target = documents[0].page_content

    # name = os.path.splitext(os.path.basename(file_path))[0]
    name = Path(file_path).stem
    
    if image_paths:
        image_iter = iter(image_paths)  # 產生 image_paths 的迭代器
    
        def figure_replacer(match):
            try:
                image_path = next(image_iter)
                return f"![Image]({image_path})"
            except StopIteration:
                return "![Image](image_missing.png)"  # 若圖片不足，放預設佔位圖

        # 將每個 <figure> 替換為對應 image_paths 中的路徑
        temp = re.sub(r"<figure>.*?</figure>", figure_replacer, target, flags=re.DOTALL)

        # 清除其餘 HTML tag
        cleaned_content = re.sub(r"<[^>]*>", "", temp)
    else:
        cleaned_content = target

    # 儲存為 Markdown 檔案
    md_path = os.path.join(md_dir, name + ".md")
    with open(md_path, 'a', encoding='utf-8') as f:
        f.write(cleaned_content + "\n")


def split_to_single_pdf(page, single_page_pdf_path):
    writer = PdfWriter()
    writer.add_page(page)
    with open(single_page_pdf_path, "wb") as single_page_pdf:
        writer.write(single_page_pdf)

# === 分析 PDF 文件 ===
def analyze_pdf(pdf_path):
    # Linux
    endpoint = os.getenv("AZURE_AIDOC_ENDPOINT")
    key = os.getenv("AZURE_AIDOC_KEY")
    
    # Windows
    # endpoint = os.environ.get("AZURE_AIDOC_ENDPOINT")
    # key = os.environ.get("AZURE_AIDOC_KEY")
    
    client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))
    with open(pdf_path, "rb") as f:
        poller = client.begin_analyze_document("prebuilt-layout", body=f)
    return poller.result()

# === 判斷是否為可疑大區塊 ===
def is_large_visual_block(polygon, azure_page):
    if isinstance(polygon[0], float):
        coords = [(polygon[i], polygon[i + 1]) for i in range(0, len(polygon), 2)]
    else:
        coords = [(p.x, p.y) for p in polygon]
    xs, ys = zip(*coords)
    width = max(xs) - min(xs)
    height = max(ys) - min(ys)
    return width > 0.2 * azure_page.width and height > 0.2 * azure_page.height

# === 將 polygon 百分比轉換為圖像座標 ===
def polygon_to_pixels(polygon, page_width, page_height, img_w, img_h):
    if isinstance(polygon[0], float):
        coords = [(polygon[i], polygon[i + 1]) for i in range(0, len(polygon), 2)]
    else:
        coords = [(p.x, p.y) for p in polygon]
    return [(x / page_width * img_w, y / page_height * img_h) for x, y in coords]

# === 主函式：抽圖 + 轉 Markdown ===
def extract_text_and_figures(pdf_path, output_img_dir, output_md_dir):
    doc = fitz.open(pdf_path)
    result = analyze_pdf(pdf_path)

    for page_idx, page in enumerate(result.pages, start=1):
        fitz_page = doc[page_idx - 1]
        matrix = fitz.Matrix(2.0, 2.0)
        pix = fitz_page.get_pixmap(matrix=matrix)
        img_data = pix.tobytes("png")
        full_image = Image.open(io.BytesIO(img_data))
        img_w, img_h = full_image.size
        azure_page = page

        page_text_blocks = []
        figure_blocks = []
        figure_polygons = set()
        
        figures = result.figures if result.figures is not None else []
        
        # === 文字區塊 ===
        for line in page.lines:
            polygon = line.polygon
            line_bbox = polygon_to_pixels(polygon, azure_page.width, azure_page.height, img_w, img_h)
            lx, ly = zip(*line_bbox)
            line_rect = (min(lx), min(ly), max(lx), max(ly))

            # 判斷是否與任何圖形區域重疊
            overlap = False
            # figures = result.figures if result.figures is not None else []
            for fig in figures:
                if fig.bounding_regions[0].page_number != page_idx:
                    continue
                fig_coords = polygon_to_pixels(fig.bounding_regions[0].polygon, azure_page.width, azure_page.height, img_w, img_h)
                fx, fy = zip(*fig_coords)
                fig_rect = (min(fx), min(fy), max(fx), max(fy))

                # 檢查兩個 bbox 是否重疊
                if not (line_rect[2] < fig_rect[0] or line_rect[0] > fig_rect[2] or
                        line_rect[3] < fig_rect[1] or line_rect[1] > fig_rect[3]):
                    overlap = True
                    break

            if not overlap:
                min_y = line_rect[1]
                page_text_blocks.append({
                    "type": "text",
                    "y": min_y,
                    "content": line.content.strip()
                })

        # === 正式圖形區塊 ===
        for fig in figures:
            if fig.bounding_regions[0].page_number != page_idx:
                continue
            polygon = fig.bounding_regions[0].polygon
            coords = polygon_to_pixels(polygon, azure_page.width, azure_page.height, img_w, img_h)
            xs, ys = zip(*coords)
            crop_box = (int(min(xs)), int(min(ys)), int(max(xs)), int(max(ys)))
            image_path = os.path.join(output_img_dir, f"{os.path.basename(pdf_path)}_page{page_idx}_fig{fig.id}.png")
            full_image.crop(crop_box).save(image_path)
            figure_blocks.append({"type": "image", "y": min(ys), "path": image_path})
            figure_polygons.add(tuple(polygon))

        # === 疑似圖形（補裁）===
        for line in page.lines:
            polygon = line.polygon
            if tuple(polygon) in figure_polygons:
                continue
            if is_large_visual_block(polygon, azure_page):
                coords = polygon_to_pixels(polygon, azure_page.width, azure_page.height, img_w, img_h)
                xs, ys = zip(*coords)
                crop_box = (int(min(xs)), int(min(ys)), int(max(xs)), int(max(ys)))
                image_path = os.path.join(output_img_dir, f"{os.path.basename(pdf_path)}_page{page_idx}_heuristic_{hash(str(polygon))}.png")
                full_image.crop(crop_box).save(image_path)
                figure_blocks.append({"type": "image", "y": min(ys), "path": image_path})

        # === 合併所有區塊並排序 ===
        all_blocks = page_text_blocks + figure_blocks
        all_blocks.sort(key=lambda b: b["y"])

        # === 輸出 Markdown ===
        output_md_path = os.path.join(output_md_dir, os.path.basename(pdf_path).replace(".pdf", ".md"))
        with open(output_md_path, "a", encoding="utf-8") as f:
            for b in all_blocks:
                if b["type"] == "text":
                    f.write(b["content"] + "\n\n")
                elif b["type"] == "image":
                    rel_path = os.path.relpath(b["path"], os.path.dirname(output_md_path)).replace("\\", "/")
                    f.write(f"![Image]({rel_path})\n\n")

def extract_figure_images(input_path, output_dir):
    
    # Linux
    endpoint = os.getenv("AZURE_AIDOC_ENDPOINT")
    key = os.getenv("AZURE_AIDOC_KEY")
    
    # Windows
    # endpoint = os.environ.get("AZURE_AIDOC_ENDPOINT")
    # key = os.environ.get("AZURE_AIDOC_KEY")
    
    document_intelligence_client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))

    with open(input_path, "rb") as f:
        poller = document_intelligence_client.begin_analyze_document(
            "prebuilt-layout",
            body=f,
            output=[AnalyzeOutputOption.FIGURES],
        )
    result: AnalyzeResult = poller.result()
    operation_id = poller.details["operation_id"]
    
    image_paths = []
    # base_name = os.path.basename(input_path).replace(".pdf", "")
    base_name = Path(input_path).stem
    if result.figures:
        for figure in result.figures:
            if figure.id:
                response = document_intelligence_client.get_analyze_result_figure(
                    model_id=result.model_id, result_id=operation_id, figure_id=figure.id
                )
                with open(os.path.join(output_dir, f"{base_name}_{figure.id}.png"), "wb") as writer:
                    writer.writelines(response)
                image_paths.append(os.path.join(output_dir, f"{base_name}_{figure.id}.png"))
        return image_paths
    else:
        print("No figures found.")

def text_image_merge_wt_md(image_paths, input_file_path):
    for image_path in image_paths:
        image_markdown = f"![Image]({image_path})"
        with open(input_file_path, "a", encoding="utf-8") as md_file:
            md_file.write("\n" + image_markdown + "\n")

def process_pdf_file(pdf_file_path):
    filename = os.path.basename(pdf_file_path)
    reader = PdfReader(pdf_file_path)
    for page_number, page in tqdm(enumerate(reader.pages, start=1), total=len(reader.pages), desc="Processing pages"):
        single_page_pdf_path = os.path.join(SPLIT_DIR, f"{os.path.splitext(filename)[0]}_page_{page_number}.pdf")
        split_to_single_pdf(page, single_page_pdf_path)
        extract_text_and_figures(single_page_pdf_path, IMG_DIR, MD_DIR)

def process_image_file(file_path):
    image_paths = extract_figure_images(file_path, IMG_DIR)
    parse_to_md(MD_DIR, file_path, image_paths)
    

def process_csv_file(csv_path):
    filename = os.path.basename(csv_path)
    df = pd.read_csv(csv_path)
    md_table = df.to_markdown(index=False)
    with open(os.path.join(MD_DIR, os.path.splitext(filename)[0] + ".md"), "w", encoding="utf-8") as f:
        f.write(f"# {filename}\n\n{md_table}\n")

def extract_excel_charts_as_images(excel_path, output_dir):

    os.makedirs(output_dir, exist_ok=True)

    app = xw.App(visible=False)
    wb = app.books.open(excel_path)
    chart_paths = []

    temp_sheet = wb.sheets.add(name="__temp_chart_export__")

    for sheet in wb.sheets:
        if sheet.name == "__temp_chart_export__":
            continue

        chart_count = 0
        for shape in sheet.shapes:
            try:
                if shape.name.lower().startswith("chart") or shape.api.Type == 3:
                    chart_count += 1
                    img_name = f"{sheet.name}_chart_{chart_count}.png"
                    image_path = os.path.join(output_dir, img_name)

                    # 複製圖表
                    shape.api.Copy()
                    time.sleep(2)  # 等待剪貼簿穩定
                    
                    # 貼上到暫存工作表
                    temp_sheet.activate()
                    temp_sheet.range("A1").select()
                    temp_sheet.api.Paste()

                    # 等待 Excel 貼上成功
                    time.sleep(0.5)

                    # 抓到貼上的圖形（最後一個 shape）
                    pasted_shape = temp_sheet.shapes[-1]
                    pasted_shape.api.Copy()

                    # 儲存為圖片（從剪貼簿貼到圖檔）
                    temp_img_path = os.path.join(output_dir, f"~tmp_{img_name}")
                    pasted_shape.api.Chart.Export(temp_img_path, "PNG")

                    # 確保儲存成功再改名
                    if os.path.exists(temp_img_path) and os.path.getsize(temp_img_path) > 0:
                        os.rename(temp_img_path, image_path)
                        chart_paths.append(image_path)
                        print(f"匯出成功：{image_path}")
                    else:
                        print(f"匯出失敗：{image_path}")

            except Exception as e:
                print(f"匯出圖表錯誤：{e}")

    # 清除暫存工作表
    try:
        temp_sheet.delete()
    except:
        pass

    wb.close()
    app.quit()
    return chart_paths

def process_xlsx_file(xlsx_path):
    # os.makedirs(MD_DIR, exist_ok=True)
    # os.makedirs(IMG_DIR, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    base_name = Path(xlsx_path).stem
    
    chart_images = extract_excel_charts_as_images(xlsx_path, IMG_DIR)
    chart_image_index = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.DataFrame(ws.values)
        if df.dropna(how='all').empty:
            continue

        df.columns = df.iloc[0]
        df = df[1:]

        md_lines = [f"# 工作表：{sheet_name}\n", df.to_markdown(index=False), "\n"]

        # 插入對應圖表圖片
        while chart_image_index < len(chart_images) and f"{sheet_name}_chart_" in os.path.basename(chart_images[chart_image_index]):
            rel_path = os.path.relpath(chart_images[chart_image_index], MD_DIR).replace(os.sep, '/')
            md_lines.append(f"![Image]({rel_path})\n")
            chart_image_index += 1

        md_filename = f"{base_name}_{sheet_name}.md"
        with open(os.path.join(MD_DIR, md_filename), "w", encoding="utf-8") as f:
            f.write("\n".join(md_lines))

def convert_docx_to_pdf(input_path, output_path=None):

    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False

    try:
        doc = word.Documents.Open(input_path)

        if output_path is None:
            output_path = os.path.splitext(input_path)[0] + ".pdf"

        doc.SaveAs(output_path, FileFormat=17)  # 17 = PDF 格式
        doc.Close()
        return output_path
    except Exception as e:
        print(f"發生錯誤：{e}")
    finally:
        word.Quit()

def process_docx_file(docx_path):
    word2pdf_path = convert_docx_to_pdf(docx_path)
    process_pdf_file(word2pdf_path)

def process_file(file_path):
    # os.makedirs(SPLIT_DIR, exist_ok=True)
    # os.makedirs(IMG_DIR, exist_ok=True)
    # os.makedirs(MD_DIR, exist_ok=True)

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        process_pdf_file(file_path)
    elif ext in [".jpg", ".jpeg", ".png"]:
        process_image_file(file_path)
    elif ext == ".csv":
        process_csv_file(file_path)
    elif ext in [".xlsx", ".xls"]:
        process_xlsx_file(file_path)
    elif ext == ".docx":
        process_docx_file(file_path)
    else:
        print(f"Unsupported file type: {ext}")

# def merge_markdown_files(input_folder, output_dir, file_name):
#     """
#     Merge Markdown (.md) files in numerical order (e.g., 'page_1', 'page_2') into a single Markdown file.

#     Args:
#         input_folder (str): The folder containing the Markdown files to merge.
#         output_file (str): The path to the output Markdown file.
#         file_name (str): The name of merged markdown file
#     """
    
#     try:
#         # Get a list of all .md files in the input folder
#         md_files = glob(os.path.join(input_folder, f"{file_name}*"))
        
#         # Sort files based on the numerical value in their filenames (e.g., 'page_1', 'page_2')
#         md_files.sort(key=lambda x: int(x.split('_')[-1].split('.')[0]))
        
#         output_file_path = os.path.join(output_dir, file_name + ".md")
        
#         # for file_path in range(md_files):
#         for file_path in tqdm(md_files):    
#             with open(output_file_path, 'a', encoding='utf-8') as outfile:
#                 with open(os.path.join(input_folder, file_path), 'r', encoding='utf-8') as infile:
#                     content = infile.read()
#                     # Write the content of the current file to the output file
#                     outfile.write(content)
#                     outfile.write("\n")  # Add spacing between files

#         print(f"Successfully merged Markdown files into '{output_dir}'.")

#     except Exception as e:
#         print(f"An error occurred: {e}")

def get_sort_key(filename):
    """
    從檔名中提取頁碼，否則回傳 float('inf') 排在最後
    """
    basename = os.path.basename(filename)
    match = re.search(r'_page_(\d+)', basename)
    if match:
        return int(match.group(1))
    else:
        return float('inf')  # 無頁碼的排在最後

def merge_markdown_files(input_folder, output_dir, file_name):
    try:
        # 找出所有以 file_name 為開頭的 .md 檔案
        md_files = glob(os.path.join(input_folder, f"{file_name}*.md"))
        
        # 依據頁碼排序，無頁碼的排最後
        md_files.sort(key=get_sort_key)

        output_file_path = os.path.join(output_dir, file_name + ".md")

        with open(output_file_path, 'w', encoding='utf-8') as outfile:
            for file_path in tqdm(md_files, desc="Merging markdown files"):
                with open(file_path, 'r', encoding='utf-8') as infile:
                    content = infile.read()
                    outfile.write(content)
                    outfile.write("\n")

        print(f"✅ Successfully merged Markdown files into '{output_file_path}'.")

    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    
    # 定義本地資料夾路徑
    ROOT_DIR = r"D:\Work\TEST"
    INPUT_DIR = os.path.join(ROOT_DIR, "input")
    SPLIT_DIR = os.path.join(ROOT_DIR, "split_files")
    MD_DIR = os.path.join(ROOT_DIR, "md_files")
    IMG_DIR = os.path.join(ROOT_DIR, "image_files")
    OUTPUT_DIR = os.path.join(ROOT_DIR, "output_files")
    
    dirs = [SPLIT_DIR, MD_DIR, IMG_DIR, OUTPUT_DIR]
    for _dir in dirs:
        os.makedirs(_dir, exist_ok=True)

    for filename in os.listdir(INPUT_DIR):
        file_path = os.path.join(INPUT_DIR, filename)

        if os.path.isfile(file_path):
            print(f"處理檔案: {file_path}")
            process_file(file_path)

        file_name = Path(file_path).stem
        merge_markdown_files(MD_DIR, OUTPUT_DIR, file_name)